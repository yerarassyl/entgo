import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


JSON_PATH = Path(r"C:\Users\erasy\AppData\Local\Temp\entgo_database_full.json")
XLSX_PATH = Path(r"C:\Users\erasy\AppData\Local\Temp\entgo_database_full.xlsx")


def q(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def jb(value):
    return q(json.dumps(value, ensure_ascii=False, separators=(",", ":"))) + "::jsonb"


def slug(value):
    value = re.sub(r"[^a-zA-Z0-9а-яА-ЯёЁ]+", "_", value).strip("_").lower()
    return value[:90] or "topic"


with JSON_PATH.open(encoding="utf-8") as f:
    data = json.load(f)

wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
sheet = wb.worksheets[2]
sheet_topics = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] and row[3]:
        sheet_topics.setdefault(row[0], []).append((str(row[3]), int(row[4] or 2)))

index = int(sys.argv[1])
subject = data["subjects"][index]
subject_id = subject["id"]
subject_name = subject["name_ru"]
lines = [
    "BEGIN;",
    f'INSERT INTO "Subject" ("id","slug","titleRu","titleKk","isRequired") VALUES ({q(subject_id)},{q(subject_id)},{q(subject_name)},{q(subject["name_kz"])},{str(subject["type"] == "mandatory").lower()}) ON CONFLICT ("id") DO UPDATE SET "titleRu"=EXCLUDED."titleRu","titleKk"=EXCLUDED."titleKk","isRequired"=EXCLUDED."isRequired";',
]

topic_ids = {}
for topic_index, (title, difficulty) in enumerate(sheet_topics.get(subject_name, []), 1):
    topic_id = f"topic_{subject_id}_sheet_{topic_index}"
    topic_ids.setdefault(title, topic_id)
    lines.append(
        f'INSERT INTO "Topic" ("id","subjectId","slug","titleRu","titleKk","difficulty","status") VALUES ({q(topic_id)},{q(subject_id)},{q("sheet_" + str(topic_index))},{q(title)},{q(title)},{difficulty},\'PUBLISHED\') ON CONFLICT ("id") DO NOTHING;'
    )

for item in subject["questions"]:
    topic_key = item["topic"]
    topic_id = f"topic_{subject_id}_{topic_key}"
    topic_ids[topic_key] = topic_id
    lines.append(
        f'INSERT INTO "Topic" ("id","subjectId","slug","titleRu","titleKk","difficulty","status") VALUES ({q(topic_id)},{q(subject_id)},{q(topic_key)},{q(topic_key)},{q(topic_key)},{int(item.get("difficulty") or 2)},\'PUBLISHED\') ON CONFLICT ("id") DO NOTHING;'
    )

for lesson in data["lessons"]:
    if lesson.get("subject_id") != subject_id:
        continue
    lesson_topic_key = lesson["topic_id"]
    topic_id = f"topic_{subject_id}_{lesson_topic_key}"
    topic_ids[lesson_topic_key] = topic_id
    lines.append(
        f'INSERT INTO "Topic" ("id","subjectId","slug","titleRu","titleKk","difficulty","status") VALUES ({q(topic_id)},{q(subject_id)},{q(lesson_topic_key)},{q(lesson_topic_key)},{q(lesson_topic_key)},2,\'PUBLISHED\') ON CONFLICT ("id") DO NOTHING;'
    )

for item in subject["questions"]:
    question_id = item["id"]
    topic_id = topic_ids[item["topic"]]
    lines.append(
        f'INSERT INTO "Question" ("id","slug","subjectId","topicId","status","locale","difficulty","body","explanation","source","sourceYear","updatedAt") VALUES ({q(question_id)},{q(question_id)},{q(subject_id)},{q(topic_id)},\'PUBLISHED\',\'RU\',{int(item.get("difficulty") or 2)},{jb({"text": item["question"]})},{jb({"text": item["explanation"]})},{q(data["meta"]["source"])},2026,CURRENT_TIMESTAMP) ON CONFLICT ("id") DO UPDATE SET "topicId"=EXCLUDED."topicId","body"=EXCLUDED."body","explanation"=EXCLUDED."explanation","status"=EXCLUDED."status","updatedAt"=CURRENT_TIMESTAMP;'
    )
    for position, option in enumerate(item["options"]):
        content = re.sub(r"^[А-ЯA-Z]\)\s*", "", str(option))
        is_correct = chr(1040 + position) == item["answer"]
        lines.append(
            f'INSERT INTO "QuestionOption" ("id","questionId","position","content","isCorrect") VALUES ({q(question_id + "_opt_" + str(position))},{q(question_id)},{position},{jb({"text": content})},{str(is_correct).lower()}) ON CONFLICT ("id") DO UPDATE SET "content"=EXCLUDED."content","isCorrect"=EXCLUDED."isCorrect";'
        )

for lesson in data["lessons"]:
    if lesson.get("subject_id") != subject_id:
        continue
    topic_id = topic_ids[lesson["topic_id"]]
    lesson_rule = "\n".join(lesson.get("key_facts", []))
    lesson_example = "\n".join(lesson.get("examples", []))
    lines.append(
        f'INSERT INTO "Lesson" ("id","topicId","summary","rule","example","mistake","steps","contentRu","publishedAt","updatedAt") VALUES ({q(lesson["id"])},{q(topic_id)},{q(lesson.get("theory", ""))},{q(lesson_rule)},{q(lesson_example)},{q("")},{jb(lesson.get("key_facts", []))},{jb(lesson)},CURRENT_TIMESTAMP,CURRENT_TIMESTAMP) ON CONFLICT ("id") DO UPDATE SET "topicId"=EXCLUDED."topicId","summary"=EXCLUDED."summary","contentRu"=EXCLUDED."contentRu","updatedAt"=CURRENT_TIMESTAMP;'
    )

lines.append("COMMIT;")
if len(sys.argv) > 2 and sys.argv[2] == "--count":
    print(len(lines))
elif len(sys.argv) > 2:
    chunk = int(sys.argv[2])
    chunk_size = 20
    start = chunk * chunk_size
    selected = lines[start:start + chunk_size]
    if selected and selected[0] == "BEGIN;":
        selected = selected[1:]
    if selected and selected[-1] == "COMMIT;":
        selected = selected[:-1]
    if selected:
        print("BEGIN;\n" + "\n".join(selected) + "\nCOMMIT;")
else:
    print("\n".join(lines))
